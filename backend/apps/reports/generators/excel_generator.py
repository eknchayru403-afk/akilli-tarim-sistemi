"""
Excel Rapor Uretici — openpyxl ile tarimsal rapor.
"""
import io
import logging
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

logger = logging.getLogger(__name__)

HEADER_FILL = PatternFill(start_color='2D6A4F', end_color='2D6A4F', fill_type='solid')
HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
ALT_FILL = PatternFill(start_color='F0F7F4', end_color='F0F7F4', fill_type='solid')
TITLE_FONT = Font(name='Calibri', size=14, bold=True, color='2D6A4F')
THIN_BORDER = Border(
    left=Side(style='thin', color='CCCCCC'), right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'), bottom=Side(style='thin', color='CCCCCC'),
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)


def generate_excel_report(report_data: dict) -> io.BytesIO:
    """Excel raporu olusturur ve BytesIO buffer dondurur."""
    wb = Workbook()
    ws_care = wb.active
    ws_care.title = 'Bakim Gecmisi'
    _write_care_sheet(ws_care, report_data)
    ws_soil = wb.create_sheet('Toprak Analizleri')
    _write_soil_sheet(ws_soil, report_data)
    ws_recs = wb.create_sheet('Urun Onerileri')
    _write_recommendation_sheet(ws_recs, report_data)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _write_header(ws, title, report_data, start_row=1):
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    cell = ws.cell(row=start_row, column=1, value=f'ATYS - {title}')
    cell.font = TITLE_FONT
    row = start_row + 1
    for label, value in [
        ('Ciftci', report_data.get('user_name', '-')),
        ('Tarla', report_data.get('field_name', 'Tum Tarlalar')),
        ('Tarih', f"{report_data.get('date_from','-')} - {report_data.get('date_to','-')}"),
        ('Rapor', f"{datetime.now():%d.%m.%Y %H:%M}"),
    ]:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1
    return row + 1


def _write_table_header(ws, row, headers):
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = THIN_BORDER


def _style_row(ws, row, cols, idx):
    for col in range(1, cols + 1):
        c = ws.cell(row=row, column=col)
        c.border = THIN_BORDER
        c.alignment = CENTER
        if idx % 2 == 1:
            c.fill = ALT_FILL


def _write_care_sheet(ws, data):
    row = _write_header(ws, 'Bakim Gecmisi', data)
    records = data.get('care_records', [])
    headers = ['Tarla', 'Mesaj', 'Oncelik', 'Durum', 'Tarih']
    for type_label, type_key in [('Sulama','irrigation'),('Gubreleme','fertilization'),('Ilaclama','pesticide')]:
        ws.cell(row=row, column=1, value=f'-- {type_label} --').font = Font(bold=True, size=12, color='2D6A4F')
        row += 1
        _write_table_header(ws, row, headers)
        row += 1
        filtered = [r for r in records if r['type'] == type_key]
        for i, r in enumerate(filtered):
            ws.cell(row=row, column=1, value=r.get('field_name','-'))
            ws.cell(row=row, column=2, value=r.get('message','-'))
            ws.cell(row=row, column=3, value=r.get('priority','-'))
            ws.cell(row=row, column=4, value='Tamamlandi' if r.get('is_done') else 'Bekliyor')
            ws.cell(row=row, column=5, value=r.get('date','-'))
            _style_row(ws, row, 5, i)
            row += 1
        if not filtered:
            ws.cell(row=row, column=1, value='Kayit bulunmamaktadir.')
            row += 1
        row += 1
    for c, w in [('A',15),('B',40),('C',12),('D',12),('E',18)]:
        ws.column_dimensions[c].width = w


def _write_soil_sheet(ws, data):
    row = _write_header(ws, 'Toprak Analizleri', data)
    headers = ['Tarla','N','P','K','pH','Nem%','Sicaklik','Yagis','Tarih']
    _write_table_header(ws, row, headers)
    row += 1
    for i, a in enumerate(data.get('soil_analyses', [])):
        for col, key in enumerate(['field_name','nitrogen','phosphorus','potassium','ph','humidity','temperature','rainfall','date'], 1):
            ws.cell(row=row, column=col, value=a.get(key, '-'))
        _style_row(ws, row, 9, i)
        row += 1
    for c in 'ABCDEFGHI':
        ws.column_dimensions[c].width = 14


def _write_recommendation_sheet(ws, data):
    row = _write_header(ws, 'Urun Onerileri', data)
    headers = ['Urun','Guven%','Verim(kg)','Kazanc(TL)','Sira','Tarih']
    _write_table_header(ws, row, headers)
    row += 1
    for i, r in enumerate(data.get('recommendations', [])):
        for col, key in enumerate(['crop_name_tr','confidence','yield_kg','revenue_tl','rank','date'], 1):
            ws.cell(row=row, column=col, value=r.get(key, '-'))
        _style_row(ws, row, 6, i)
        row += 1
    for c in 'ABCDEF':
        ws.column_dimensions[c].width = 18
